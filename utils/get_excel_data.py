import pandas as pd

import openpyxl


# class ExcelUtils(object):
def get_data():
    data_buffer = openpyxl.load_workbook("/resources/item.xlsx")
    data = data_buffer.active
    brands_list = []
    for row in range(1, 4):
        for col in data.iter_cols(3, 3):
            brands_list.append(col[row].value)
    return brands_list

    # def __init__(self, driver):
    #     self.driver = driver
#---------------------------------------------------------------------------------------------------------
    # def get_excel_data(self):
    #     dataframe = openpyxl.load_workbook("/resources/item.xlsx")
    #     # Define variable to read sheet
    #     dataframe1 = dataframe.active
    #     brands_list = []
    #     # Iterate the loop to read the cell values
    #     for row in range(1, 4):
    #         for col in dataframe1.iter_cols(3, 3):
    #             # print(col[row].value)
    #             brands_list.append(col[row].value)
    #     # print(brands_list)
    #     return brands_list
# ---------------------------------------------------------------------------------------------------------
# class GetExcelData():

# def get_excel_data(self):
#     with pd.ExcelFile('/Users/tarasodynyuk/PycharmProjects/web_tests/resources/item.xlsx') as xls:
#         df1 = pd.read_excel(xls, 'mobileSheet')
#     print("****Result Sheet 1****")
#     print(df1[1:3]['brand'])

# data = pd.read_excel('resources/item.xlsx')
# print(data.loc[[1, 2, 3], ['brand']])
# print("nv jhhjh")

